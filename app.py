import streamlit as st
from datetime import datetime, timedelta, date, time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# ─────────────────────────────────────────────────────────────────────────────
# Configuração da página
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Relatório de Participações | Teams",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1F3864 0%, #2E75B6 100%);
        padding: 1.2rem 1.5rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
    }
    .main-header h1 { color: white; margin: 0; font-size: 1.6rem; }
    .main-header p  { color: #cde0f5; margin: 0.3rem 0 0 0; font-size: 0.9rem; }
    .section-title {
        font-size: 1rem; font-weight: 700; color: #1F3864;
        border-bottom: 2px solid #2E75B6; padding-bottom: 4px; margin: 1rem 0 0.6rem 0;
    }
    .module-card {
        background: #f0f4fb; border-left: 4px solid #2E75B6;
        border-radius: 6px; padding: 0.6rem 0.8rem; margin-bottom: 0.5rem;
    }
    .stat-box {
        background: #f8f9fa; border: 1px solid #dee2e6;
        border-radius: 8px; padding: 0.8rem; text-align: center;
    }
    .stat-box .value { font-size: 1.5rem; font-weight: 700; color: #1F3864; }
    .stat-box .label { font-size: 0.75rem; color: #6c757d; }
    .ok-row   { color: #198754; font-weight: 600; }
    .warn-row { color: #dc3545; font-weight: 600; }
    div[data-testid="stSidebarContent"] { background: #f8faff; }
    .stButton > button {
        width: 100%; background: #2E75B6; color: white;
        border: none; border-radius: 6px; padding: 0.6rem;
        font-weight: 600; font-size: 0.95rem;
    }
    .stButton > button:hover { background: #1F3864; }
    .stDownloadButton > button {
        width: 100%; background: #198754; color: white;
        border: none; border-radius: 6px; padding: 0.6rem;
        font-weight: 600; font-size: 0.95rem;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Funções auxiliares
# ─────────────────────────────────────────────────────────────────────────────
FMT_DT = "%m/%d/%y, %I:%M:%S %p"

def parse_dt(s):
    try:
        return datetime.strptime(s.strip(), FMT_DT)
    except Exception:
        return None

def fmt_dur(td):
    s = int(td.total_seconds())
    if s <= 0: return "—"
    h, r = divmod(s, 3600)
    m, sec = divmod(r, 60)
    if h > 0:   return f"{h}h {m:02d}min {sec:02d}s"
    elif m > 0: return f"{m}min {sec:02d}s"
    else:       return f"{sec}s"

def fmt_time(dt):
    return dt.strftime("%H:%M:%S")

def parse_csv(file_bytes):
    """Lê o CSV do Teams (UTF-16) e devolve dict email -> {nome, sessoes[]}"""
    try:
        content = file_bytes.decode("utf-16")
    except Exception:
        try:
            content = file_bytes.decode("utf-8")
        except Exception:
            content = file_bytes.decode("latin-1")

    lines = content.splitlines()
    activities = {}
    organizer_email = None
    in_part = False
    in_act  = False

    for line in lines:
        # Detectar organizador na secção Participantes
        if line.startswith("2. Participantes"):
            in_part = True; in_act = False; continue
        if line.startswith("3. Atividades"):
            in_act = True; in_part = False; continue
        if line.startswith("4."):
            in_act = False; in_part = False; continue

        parts = line.split("\t")

        if in_part and len(parts) >= 7:
            if parts[0] == "Nome": continue
            email  = parts[4].strip()
            funcao = parts[6].strip() if len(parts) > 6 else ""
            if funcao == "Organizador":
                organizer_email = email

        if in_act and len(parts) >= 5:
            if parts[0] == "Nome": continue
            nome, e_s, s_s, _, email = parts[0], parts[1], parts[2], parts[3], parts[4]
            e, s = parse_dt(e_s), parse_dt(s_s)
            if e and s:
                if email not in activities:
                    activities[email] = {"nome": nome.replace(" (Convidado)", ""), "sessoes": []}
                activities[email]["sessoes"].append((e, s))

    for d in activities.values():
        d["sessoes"].sort(key=lambda x: x[0])

    # Detectar data da sessão a partir do primeiro timestamp do CSV
    sessao_date = None
    all_entries = [e for d in activities.values() for e, _ in d["sessoes"]]
    if all_entries:
        sessao_date = min(all_entries).date()

    return activities, organizer_email, sessao_date

def presence_in_window(sessoes, win_ini, win_fim):
    """Calcula presença, ausências, atraso e saída antecipada numa janela temporal."""
    clipped = []
    for e, s in sessoes:
        cs, ce = max(e, win_ini), min(s, win_fim)
        if ce > cs:
            clipped.append((cs, ce))

    win_dur = win_fim - win_ini
    if not clipped:
        return timedelta(0), [(win_ini, win_fim, win_dur)], win_dur, timedelta(0)

    total        = sum((end - start for start, end in clipped), timedelta())
    atraso_td    = clipped[0][0] - win_ini   if clipped[0][0] > win_ini   else timedelta(0)
    saida_ant_td = win_fim - clipped[-1][1]  if clipped[-1][1] < win_fim  else timedelta(0)

    ausencias = []
    if atraso_td.total_seconds() > 60:
        ausencias.append((win_ini, clipped[0][0], atraso_td))
    for i in range(1, len(clipped)):
        gap = clipped[i][0] - clipped[i-1][1]
        if gap.total_seconds() > 60:
            ausencias.append((clipped[i-1][1], clipped[i][0], gap))
    if saida_ant_td.total_seconds() > 60:
        ausencias.append((clipped[-1][1], win_fim, saida_ant_td))

    return total, ausencias, atraso_td, saida_ant_td

def build_rows(activities, organizer_email, modulos):
    """Constrói a lista de linhas do relatório."""
    sessao_inicio = modulos[0]["inicio"]
    sessao_fim    = modulos[-1]["fim"]
    rows = []

    for email, data in activities.items():
        if email == organizer_email:
            continue
        sessoes = data["sessoes"]
        nome    = data["nome"]

        mod_data = []
        for mod in modulos:
            pres, aus_list, atraso, saida_ant = presence_in_window(
                sessoes, mod["inicio"], mod["fim"]
            )
            total_aus = sum((d for _, _, d in aus_list), timedelta())
            dur_mod   = mod["fim"] - mod["inicio"]
            limite    = dur_mod * 0.20  # 20% da duração do módulo
            reprovado = total_aus > limite

            obs = []
            n_mid = sum(
                1 for gi, gf, _ in aus_list
                if gi > mod["inicio"] and gf < mod["fim"]
            )
            if atraso.total_seconds()    > 60: obs.append(f"Atraso {fmt_dur(atraso)}")
            if n_mid                     > 0:  obs.append(f"{n_mid} ausência(s) durante módulo")
            if saida_ant.total_seconds() > 60: obs.append(f"Saiu {fmt_dur(saida_ant)} antes do fim")
            mod_data.append({
                "pres": pres, "aus": total_aus,
                "limite_aus": limite,
                "obs": "; ".join(obs) if obs else "—",
                "has_issue": bool(obs),
                "reprovado": reprovado,
            })

        total_pres  = sum((m["pres"] for m in mod_data), timedelta())
        total_aus   = (sessao_fim - sessao_inicio) - total_pres
        reprovado   = any(m["reprovado"] for m in mod_data)
        rows.append({
            "nome": nome, "email": email,
            "entrada": sessoes[0][0], "saida": sessoes[-1][1],
            "sessoes_list": sessoes,
            "mod_data": mod_data,
            "total_pres": total_pres, "total_aus": total_aus,
            "reprovado": reprovado,
        })

    rows.sort(key=lambda r: r["entrada"])
    return rows, sessao_inicio, sessao_fim

# ─────────────────────────────────────────────────────────────────────────────
# Geração do Excel
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel(rows, modulos, sessao_inicio, sessao_fim,
                   curso_nome, data_sessao, formador_nome):
    COR_TITULO    = "2E75B6"
    COR_OK        = "E2EFDA"
    COR_ALERT     = "FCE4D6"
    CORES_MOD_HDR = ["1F3864","375623","7B3F00","5C2D91","8B0000","00695C"]
    CORES_MOD_ROW = ["DAE3F3","D9EAD3","FFF3E0","EDE7F6","FCE4EC","E0F2F1"]
    COR_TOTAL_HDR = "3F3151"
    RED           = "C00000"

    def bd():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    def hc(ws, r, c, v, bg, ft="FFFFFF", bold=True, sz=10, wrap=True):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=bold, color=ft, size=sz)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        cell.border    = bd()

    def dc(ws, r, c, v, bg=None, bold=False, align="center", ft="000000", sz=10):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=bold, color=ft, size=sz)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border    = bd()
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)

    wb = openpyxl.Workbook()

    # ── FOLHA 1: Resumo ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo de Participações"
    ws.sheet_view.showGridLines = False

    n_mod = len(modulos)
    # Colunas: 5 identificação + 4*n_mod módulos + 4 totais (inclui resultado por módulo e global)
    total_cols = 5 + 4 * n_mod + 4
    COR_APROVADO  = "C6EFCE"
    COR_REPROVADO = "FFC7CE"
    COR_APR_FT    = "276221"
    COR_REP_FT    = "9C0006"

    def col_letter(n): return get_column_letter(n)

    # Row 1 – título principal
    ws.merge_cells(f"A1:{col_letter(total_cols)}1")
    c = ws["A1"]
    c.value      = f"{curso_nome}  —  Relatório de Participações  |  {data_sessao.strftime('%d/%m/%Y')}  |  {formador_nome}"
    c.font       = Font(bold=True, size=13, color="FFFFFF")
    c.fill       = PatternFill("solid", fgColor=COR_TITULO)
    c.alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2 – blocos de cabeçalho por módulo
    ws.merge_cells(f"A2:E2")
    c = ws["A2"]; c.value = "Identificação"
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="444444")
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = bd()

    col_offset = 6
    for mi, mod in enumerate(modulos):
        start_col = col_offset
        end_col   = col_offset + 3
        ws.merge_cells(f"{col_letter(start_col)}2:{col_letter(end_col)}2")
        c = ws[f"{col_letter(start_col)}2"]
        dur_mod = mod["fim"] - mod["inicio"]
        limite  = dur_mod * 0.20
        c.value = (f"Módulo {mi+1} — {mod['nome']}  "
                   f"({mod['inicio'].strftime('%H:%M')} – {mod['fim'].strftime('%H:%M')})  "
                   f"| Limite ausência: {fmt_dur(limite)}")
        c.font      = Font(bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=CORES_MOD_HDR[mi % len(CORES_MOD_HDR)])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bd()
        col_offset += 4

    ws.merge_cells(f"{col_letter(col_offset)}2:{col_letter(col_offset+3)}2")
    c = ws[f"{col_letter(col_offset)}2"]
    c.value = f"Totais da Sessão  ({sessao_inicio.strftime('%H:%M')} – {sessao_fim.strftime('%H:%M')})"
    c.font      = Font(bold=True, size=10, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=COR_TOTAL_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = bd()
    ws.row_dimensions[2].height = 22

    # Row 3 – cabeçalhos de colunas
    hdrs = [("Nº","444444"),("Nome","444444"),("E-mail","444444"),
            ("1ª Entrada","444444"),("Última\nSaída","444444")]
    for mi, mod in enumerate(modulos):
        bg = CORES_MOD_HDR[mi % len(CORES_MOD_HDR)]
        hdrs += [(f"Presente\nMód.{mi+1}", bg),
                 (f"Ausente\nMód.{mi+1}",  bg),
                 (f"Ocorrências\nMód.{mi+1}", bg),
                 (f"Resultado\nMód.{mi+1}", bg)]
    hdrs += [("Total\nPresente",   COR_TOTAL_HDR),
             ("Total\nAusente",   COR_TOTAL_HDR),
             ("Síntese",          COR_TOTAL_HDR),
             ("RESULTADO\nFINAL", "880000")]
    for col, (h, bg) in enumerate(hdrs, 1):
        hc(ws, 3, col, h, bg)
    ws.row_dimensions[3].height = 40

    # Linhas de dados
    for i, r in enumerate(rows):
        rn = i + 4
        any_issue = any(m["has_issue"] for m in r["mod_data"])
        row_bg    = COR_ALERT if any_issue else COR_OK

        syn_parts = [f"Mod.{j+1}: {m['obs']}"
                     for j, m in enumerate(r["mod_data"]) if m["obs"] != "—"]
        syn = "; ".join(syn_parts) if syn_parts else "Sem ocorrências"

        dc(ws, rn,  1, i+1,                      bg=row_bg)
        dc(ws, rn,  2, r["nome"], bold=True,      bg=row_bg, align="left")
        dc(ws, rn,  3, r["email"],                bg=row_bg, align="left")
        dc(ws, rn,  4, fmt_time(r["entrada"]),    bg=row_bg)
        dc(ws, rn,  5, fmt_time(r["saida"]),      bg=row_bg)

        col_offset = 6
        for mi, m in enumerate(r["mod_data"]):
            m_bg = ("F4CCCC" if m["has_issue"]
                    else CORES_MOD_ROW[mi % len(CORES_MOD_ROW)])
            aus_s = fmt_dur(m["aus"]) if m["aus"].total_seconds() > 60 else "—"
            res_bg = COR_REPROVADO if m["reprovado"] else COR_APROVADO
            res_ft = COR_REP_FT   if m["reprovado"] else COR_APR_FT
            res_v  = "REPROVADO" if m["reprovado"] else "APROVADO"
            dc(ws, rn, col_offset,   fmt_dur(m["pres"]),  bg=m_bg, bold=True)
            dc(ws, rn, col_offset+1, aus_s, bg=m_bg,
               bold=m["aus"].total_seconds()>60,
               ft=RED if m["aus"].total_seconds()>60 else "000000")
            dc(ws, rn, col_offset+2, m["obs"], bg=m_bg, align="left",
               ft=RED if m["has_issue"] else "000000")
            dc(ws, rn, col_offset+3, res_v, bg=res_bg, bold=True, ft=res_ft)
            col_offset += 4

        t_aus_s  = fmt_dur(r["total_aus"]) if r["total_aus"].total_seconds()>60 else "—"
        res_glob_bg = COR_REPROVADO if r["reprovado"] else COR_APROVADO
        res_glob_ft = COR_REP_FT   if r["reprovado"] else COR_APR_FT
        res_glob_v  = "REPROVADO"  if r["reprovado"] else "APROVADO"
        dc(ws, rn, col_offset,   fmt_dur(r["total_pres"]), bg=row_bg, bold=True)
        dc(ws, rn, col_offset+1, t_aus_s, bg=row_bg,
           bold=r["total_aus"].total_seconds()>60,
           ft=RED if r["total_aus"].total_seconds()>60 else "000000")
        dc(ws, rn, col_offset+2, syn, bg=row_bg, align="left",
           ft=RED if any_issue else "000000")
        dc(ws, rn, col_offset+3, res_glob_v, bg=res_glob_bg, bold=True,
           ft=res_glob_ft, sz=11)
        ws.row_dimensions[rn].height = 20

    # Legenda
    lr = len(rows) + 5
    ws.merge_cells(f"A{lr}:{col_letter(total_cols)}{lr}")
    c = ws[f"A{lr}"]
    c.value = ("Legenda:  🟢 Verde = sem ocorrências   🟠 Laranja = ocorrências   "
               "✅ APROVADO = ausência ≤ 20% por módulo   ❌ REPROVADO = ausência > 20% em algum módulo")
    c.font      = Font(italic=True, size=9, color="444444")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[lr].height = 15

    # Larguras
    widths = [5, 26, 34, 11, 11]
    for _ in modulos:
        widths += [15, 15, 36, 14]
    widths += [15, 15, 44, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[col_letter(col)].width = w

    # ── FOLHA 2: Detalhe de Sessões ───────────────────────────────────────────
    ws2 = wb.create_sheet("Detalhe de Sessões")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    c = ws2["A1"]
    c.value = (f"Detalhe de Sessões e Ausências  |  {curso_nome}  "
               f"|  {data_sessao.strftime('%d/%m/%Y')}"
               f"  ({sessao_inicio.strftime('%H:%M')} – {sessao_fim.strftime('%H:%M')})")
    c.font      = Font(bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=COR_TITULO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    for col, h in enumerate(["Nome / Username","Sessão","Hora de Entrada",
                              "Hora de Saída","Duração Sessão",
                              "Intervalo até próxima sessão","Duração do Intervalo"], 1):
        hc(ws2, 2, col, h, "1F3864")
    ws2.row_dimensions[2].height = 36

    row2 = 3
    for r in rows:
        sessoes_list = r["sessoes_list"]
        nome = r["nome"]
        for idx, (entrada, saida) in enumerate(sessoes_list):
            dur = saida - entrada
            if idx < len(sessoes_list) - 1:
                gi, gf = saida, sessoes_list[idx+1][0]
                gd_in  = min(gf, sessao_fim) - max(gi, sessao_inicio)
                show   = gd_in.total_seconds() > 60
                gap_label = f"{fmt_time(gi)} → {fmt_time(gf)}" if show else "—"
                gap_str   = fmt_dur(gd_in) if show else "—"
                bg = "FFF2CC" if show else "DAE3F3"
            else:
                gap_label, gap_str, bg = "—", "—", COR_OK

            dc(ws2, row2, 1, nome if idx == 0 else "", bold=(idx==0), bg=bg, align="left")
            dc(ws2, row2, 2, f"Sessão {idx+1}", bg=bg)
            dc(ws2, row2, 3, fmt_time(entrada), bg=bg)
            dc(ws2, row2, 4, fmt_time(saida),   bg=bg)
            dc(ws2, row2, 5, fmt_dur(dur),       bg=bg)
            dc(ws2, row2, 6, gap_label, bg=bg,
               ft="C00000" if gap_str != "—" else "000000")
            dc(ws2, row2, 7, gap_str, bold=(gap_str != "—"), bg=bg,
               ft="C00000" if gap_str != "—" else "000000")
            ws2.row_dimensions[row2].height = 17
            row2 += 1

        for col in range(1, 8):
            c = ws2.cell(row=row2, column=col)
            c.fill   = PatternFill("solid", fgColor="F2F2F2")
            c.border = bd()
        ws2.row_dimensions[row2].height = 5
        row2 += 1

    for col, w in enumerate([28, 12, 16, 16, 16, 30, 20], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────────────────────────────────────
# Estado da sessão
# ─────────────────────────────────────────────────────────────────────────────
if "modulos" not in st.session_state:
    st.session_state.modulos = [
        {"nome": "Módulo 1", "hora_ini": time(19, 0), "hora_fim": time(21, 0)},
    ]

def add_modulo():
    ultimo = st.session_state.modulos[-1]
    st.session_state.modulos.append({
        "nome": f"Módulo {len(st.session_state.modulos)+1}",
        "hora_ini": ultimo["hora_fim"],
        "hora_fim": ultimo["hora_fim"],
    })

def remove_modulo(idx):
    if len(st.session_state.modulos) > 1:
        st.session_state.modulos.pop(idx)

# ─────────────────────────────────────────────────────────────────────────────
# Interface
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <h1>📋 Relatório de Participações — Microsoft Teams</h1>
  <p>Carregue o ficheiro CSV exportado do Teams, configure o cronograma e gere o relatório Excel.</p>
</div>
""", unsafe_allow_html=True)

# ── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuração da Sessão")

    st.markdown('<div class="section-title">📁 Ficheiro Teams</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "Relatório de participações (.csv)",
        type=["csv"],
        help="Exportado diretamente do Microsoft Teams → Participantes → Transferir lista de participantes",
    )

    # Pré-processar CSV para detetar data automaticamente
    if uploaded:
        try:
            file_bytes_preview = uploaded.getvalue()  # lê sem consumir o cursor
            _, _, detected_date = parse_csv(file_bytes_preview)
            if detected_date and st.session_state.get("last_file") != uploaded.name:
                st.session_state["detected_date"] = detected_date
                st.session_state["last_file"] = uploaded.name
        except Exception:
            pass

    st.markdown('<div class="section-title">📚 Informação do Curso</div>', unsafe_allow_html=True)
    curso_nome  = st.text_input("Nome / Número do Curso", value="EF278_TCC_FI_AF (nº11866)")
    default_date = st.session_state.get("detected_date", date.today())
    if uploaded:
        st.caption(f"📅 Data detetada no CSV: **{default_date.strftime('%d/%m/%Y')}**")
    data_sessao = st.date_input("Data da Sessão", value=default_date)
    formador    = st.text_input("Formador(a)", value="")

    st.markdown('<div class="section-title">🕐 Módulos da Sessão</div>', unsafe_allow_html=True)
    st.caption("Defina os módulos por ordem cronológica. O fim de um deve coincidir com o início do seguinte.")

    for i, mod in enumerate(st.session_state.modulos):
        with st.container():
            st.markdown(f'<div class="module-card"><strong>Módulo {i+1}</strong></div>',
                        unsafe_allow_html=True)
            c1, c2 = st.columns([3, 1])
            with c1:
                mod["nome"] = st.text_input(
                    "Nome do módulo", value=mod["nome"],
                    key=f"mod_nome_{i}", label_visibility="collapsed"
                )
            with c2:
                if st.button("✕", key=f"del_{i}", help="Remover módulo",
                             disabled=len(st.session_state.modulos) == 1):
                    remove_modulo(i)
                    st.rerun()

            c3, c4 = st.columns(2)
            with c3:
                mod["hora_ini"] = st.time_input("Início", value=mod["hora_ini"],
                                                key=f"ini_{i}", step=1800)
            with c4:
                mod["hora_fim"] = st.time_input("Fim",    value=mod["hora_fim"],
                                                key=f"fim_{i}", step=1800)

    if st.button("＋ Adicionar Módulo"):
        add_modulo()
        st.rerun()

    st.divider()
    gerar_btn = st.button("🚀 Gerar Relatório", type="primary")

# ── ÁREA PRINCIPAL ────────────────────────────────────────────────────────────
if not uploaded:
    st.info("👈 Comece por carregar o ficheiro CSV do Teams na barra lateral.")

    with st.expander("ℹ️ Como exportar o relatório de participações do Teams?", expanded=True):
        st.markdown("""
1. Durante ou após a reunião, clique em **Participantes** (ícone de pessoas)
2. No painel, clique em **⋯ Mais opções** → **Transferir lista de participantes**
3. Será descarregado um ficheiro `.csv` — carregue-o aqui
        """)

    with st.expander("📋 Sobre a configuração de módulos"):
        st.markdown("""
- Cada **módulo** corresponde a um bloco temático com hora de início e fim próprios
- Uma sessão pode ter **um ou vários módulos** encadeados
- O relatório calculará presença, atrasos e ausências separadamente para cada módulo
- Exemplo: Módulo 1 (19h–21h) + Módulo 2 (21h–23h)
        """)
else:
    # Processar CSV
    file_bytes = uploaded.getvalue()
    try:
        activities, organizer_email, _ = parse_csv(file_bytes)
    except Exception as e:
        st.error(f"Erro ao processar o ficheiro: {e}")
        st.stop()

    # Construir módulos com datas
    modulos_config = []
    valid = True
    for mod in st.session_state.modulos:
        ini = datetime.combine(data_sessao, mod["hora_ini"])
        fim = datetime.combine(data_sessao, mod["hora_fim"])
        if fim <= ini:
            st.error(f"⚠️ '{mod['nome']}': a hora de fim deve ser posterior à de início.")
            valid = False
        modulos_config.append({"nome": mod["nome"], "inicio": ini, "fim": fim})

    if not valid:
        st.stop()

    # Calcular dados
    rows, sessao_ini, sessao_fim = build_rows(activities, organizer_email, modulos_config)
    n_formandos  = len(rows)
    n_ocorrencias = sum(1 for r in rows if any(m["has_issue"] for m in r["mod_data"]))
    n_ok          = n_formandos - n_ocorrencias
    dur_sessao    = sessao_fim - sessao_ini

    # ── Estatísticas rápidas ──────────────────────────────────────────────────
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Formandos", n_formandos)
    with col2:
        st.metric("Sem ocorrências", n_ok,
                  delta=f"{round(n_ok/n_formandos*100)}%" if n_formandos else None,
                  delta_color="normal")
    with col3:
        st.metric("Com ocorrências", n_ocorrencias,
                  delta=f"{round(n_ocorrencias/n_formandos*100)}%" if n_formandos else None,
                  delta_color="inverse")
    with col4:
        st.metric("Duração da sessão", fmt_dur(dur_sessao))

    st.divider()

    # ── Pré-visualização ──────────────────────────────────────────────────────
    st.markdown("### 👥 Pré-visualização dos Formandos")

    tab1, tab2 = st.tabs(["📊 Resumo por Módulo", "📋 Detalhe de Sessões"])

    with tab1:
        table_data = []
        for r in rows:
            row_d = {
                "Nome": r["nome"],
                "1ª Entrada": fmt_time(r["entrada"]),
                "Última Saída": fmt_time(r["saida"]),
            }
            for mi, m in enumerate(r["mod_data"]):
                row_d[f"Mod.{mi+1} Presente"]    = fmt_dur(m["pres"])
                row_d[f"Mod.{mi+1} Ausente"]     = fmt_dur(m["aus"]) if m["aus"].total_seconds() > 60 else "—"
                row_d[f"Mod.{mi+1} Ocorrências"] = m["obs"]
                row_d[f"Mod.{mi+1} Resultado"]   = "❌ REPROVADO" if m["reprovado"] else "✅ APROVADO"
            row_d["Total Presente"] = fmt_dur(r["total_pres"])
            row_d["Total Ausente"]  = fmt_dur(r["total_aus"]) if r["total_aus"].total_seconds()>60 else "—"
            row_d["Resultado Final"] = "❌ REPROVADO" if r["reprovado"] else "✅ APROVADO"
            table_data.append(row_d)

        if table_data:
            import pandas as pd
            df = pd.DataFrame(table_data)

            resultado_cols = [c for c in df.columns if "Resultado" in c]

            def color_resultado(val):
                if "REPROVADO" in str(val): return "background-color: #FFC7CE; color: #9C0006; font-weight:bold"
                if "APROVADO"  in str(val): return "background-color: #C6EFCE; color: #276221; font-weight:bold"
                return ""

            styled = df.style.map(color_resultado, subset=resultado_cols)
            st.dataframe(styled, use_container_width=True, hide_index=True)

            # Resumo rápido de aprovações
            n_rep = sum(1 for r in rows if r["reprovado"])
            n_apr = len(rows) - n_rep
            ca, cb = st.columns(2)
            ca.success(f"✅ **{n_apr} Aprovados**")
            if n_rep > 0:
                cb.error(f"❌ **{n_rep} Reprovados** — ausência > 20% em algum módulo")

    with tab2:
        for r in rows:
            has_issue = any(m["has_issue"] for m in r["mod_data"])
            icon = "⚠️" if has_issue else "✅"
            with st.expander(f"{icon} {r['nome']}  —  {fmt_time(r['entrada'])} → {fmt_time(r['saida'])}"):
                for idx, (entrada, saida) in enumerate(r["sessoes_list"]):
                    dur = saida - entrada
                    col_a, col_b, col_c = st.columns([2, 2, 2])
                    col_a.markdown(f"**Sessão {idx+1}**")
                    col_b.markdown(f"🟢 `{fmt_time(entrada)}` → `{fmt_time(saida)}`")
                    col_c.markdown(f"⏱ {fmt_dur(dur)}")

                    if idx < len(r["sessoes_list"]) - 1:
                        gi, gf = saida, r["sessoes_list"][idx+1][0]
                        gd = gf - gi
                        if gd.total_seconds() > 60:
                            st.markdown(
                                f"&nbsp;&nbsp;&nbsp;&nbsp;🔴 **Ausência:** "
                                f"`{fmt_time(gi)}` → `{fmt_time(gf)}`  —  **{fmt_dur(gd)}**"
                            )

    # ── Gerar Excel ───────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 📥 Gerar Relatório Excel")

    if gerar_btn or st.button("📊 Gerar e Transferir Excel"):
        with st.spinner("A gerar o relatório..."):
            buf = generate_excel(
                rows, modulos_config, sessao_ini, sessao_fim,
                curso_nome, data_sessao, formador
            )
        nome_ficheiro = (
            f"Relatorio_Participacoes_{curso_nome.replace(' ','_').replace('/','_')}"
            f"_{data_sessao.strftime('%Y%m%d')}.xlsx"
        )
        st.success("✅ Relatório gerado com sucesso!")
        st.download_button(
            label="⬇️ Transferir Excel",
            data=buf,
            file_name=nome_ficheiro,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
